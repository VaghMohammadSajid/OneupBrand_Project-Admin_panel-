from django.shortcuts import render
from useraccount.models import (
    IntermediateProductTableChild,
    IntermediateProductTableError,
    IntermediateProductTableSucces,
    UploadProductJSon,
    ProductUploadErrorLog,
    ProductUploadSuccesLog,
)
from django.shortcuts import render, redirect
from oscar.core.loading import get_model
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import tempfile
from django.core.paginator import Paginator

from django.conf import settings
import os

# Create your views here.
Product = get_model("catalogue", "Product")


def product_error_view(request):
    log_list = ProductUploadErrorLog.objects.all().order_by("-upload_date")
    paginator = Paginator(log_list, 100)

    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        template_name="report/product_error_log.html",
        context={"log": page_obj},
    )


def download_error_log(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Product Error Log"

    headers = ["S.No", "Upload Date", "Upload Count"]
    worksheet.append(headers)

    logs = ProductUploadErrorLog.objects.all().order_by("-upload_date")

    for index, log in enumerate(logs, start=1):
        worksheet.append(
            [index, log.upload_date.strftime("%d/%m/%Y %H:%M"), log.upload_count]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=product_error_log.xlsx"

    workbook.save(response)

    return response


def download_success_log(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Product Success Log"

    headers = ["S.No", "Upload Date", "Upload Count"]
    worksheet.append(headers)

    logs = ProductUploadSuccesLog.objects.all().order_by("-upload_date")

    for index, log in enumerate(logs, start=1):
        worksheet.append(
            [index, log.upload_date.strftime("%d/%m/%Y %H:%M"), log.upload_count]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=product_success_log.xlsx"

    workbook.save(response)

    return response


def product_success_view(request):
    log_list = ProductUploadSuccesLog.objects.all().order_by("-upload_date")
    paginator = Paginator(log_list, 100)

    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        template_name="report/product_success_log.html",
        context={"log": page_obj},
    )


def error_product_list(request, id):
    product = IntermediateProductTableError.objects.filter(log_date_conn__id=id)
    return render(
        request=request,
        template_name="report/error_product_list.html",
        context={"product": product},
    )


def success_product_list(request, id):
    product = IntermediateProductTableSucces.objects.filter(log_date_conn__id=id)
    return render(
        request=request,
        template_name="report/error_product_list.html",
        context={"product": product},
    )


def download_product_csv(request):
    product_query_set = Product.objects.all()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="yourmodel_data.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "YourModel Data"

    # Header row
    ws.append(["upc", "Image"])

    for obj in product_query_set:
        # Create a temporary directory to store images
        temp_dir = tempfile.mkdtemp()

        # Save the image to the temporary directory
        image_path = os.path.join(temp_dir, "image.png")
        with open(image_path, "wb") as img_file:
            img_file.write(obj.images.all()[0].original.read())

        # Embed the image as a drawing
        img = ExcelImage(image_path)
        img.width = 50  # Set thumbnail width
        img.height = 20  # Set thumbnail height

        # Add data to Excel
        ws.append([obj.upc])

        # Add the image as a drawing
        img.anchor = f"C{ws.max_row}"
        ws.add_image(img)

    wb.save(response)
    return response
